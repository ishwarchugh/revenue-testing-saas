from pathlib import Path

import asyncio
import datetime as dt
import io
import numbers
import json
import math
import os
import re
import base64
from uuid import uuid4

import pandas as pd
from dateutil.parser import parse as dateutil_parse
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from starlette.responses import StreamingResponse

from .revenue_tests import cutoff_testing, mus_sample_size_parameters, mus_sampling, target_testing

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

app = FastAPI(title="Revenue Testing SaaS API", version="0.1.0")

FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"
OUTPUTS_DIR = Path(__file__).resolve().parent.parent / "outputs"

# Mount the frontend folder so any static assets can be served.
app.mount("/frontend", StaticFiles(directory=str(FRONTEND_DIR)), name="frontend")


@app.get("/")
def root() -> FileResponse:
    return FileResponse(str(FRONTEND_DIR / "index.html"))


@app.get("/testing")
def testing() -> FileResponse:
    return FileResponse(str(FRONTEND_DIR / "testing.html"))


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


def _coerce_bool(value: str | None) -> bool:
    if value is None:
        return False
    v = str(value).strip().lower()
    return v in {"1", "true", "t", "yes", "y", "on"}


def _norm_inv(value: object) -> str:
    return str("" if value is None else value).strip()


def _safe_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        if isinstance(value, float) and pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, str):
        t = value.strip().replace(",", "").replace("$", "").replace("€", "").replace("£", "").strip()
        if not t:
            return None
        if re.search(r"[A-Za-z]", t):
            return None
        try:
            return float(t)
        except Exception:
            return None
    try:
        v = float(value)
        return v
    except Exception:
        return None


def _row_get_ignore_case(record: dict, header: str) -> object | None:
    want = header.strip().lower()
    for k, v in record.items():
        if str(k).strip().lower() == want:
            return v
    return None


def _row_gl_date_raw(record: dict) -> object | None:
    """First non-empty cell among GL Date / Date (case-insensitive header match)."""
    for hdr in ("GL Date", "Date"):
        v = _row_get_ignore_case(record, hdr)
        if v is None:
            continue
        try:
            if isinstance(v, float) and pd.isna(v):
                continue
        except Exception:
            pass
        if isinstance(v, str) and not v.strip():
            continue
        return v
    return None


def _gl_date_display_dd_mm_yyyy(val: object) -> str:
    """Parse workbook / Excel serial / datetime / text; return DD/MM/YYYY or '' if parsing fails."""
    print(f"[DATE DEBUG] val={val} type={type(val).__name__}")
    if val is None or val == "":
        return ""

    if isinstance(val, bool):
        return ""

    try:
        if isinstance(val, float) and pd.isna(val):
            return ""
    except Exception:
        pass

    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return ""
        try:
            return val.strftime("%d/%m/%Y")
        except Exception:
            return ""

    if isinstance(val, dt.date):
        return val.strftime("%d/%m/%Y")

    if isinstance(val, numbers.Real) and not isinstance(val, bool):
        try:
            if isinstance(val, float) and pd.isna(val):
                return ""
        except Exception:
            pass
        try:
            serial = int(val)
            base = dt.datetime(1899, 12, 30)
            parsed = base + dt.timedelta(days=serial)
            return parsed.strftime("%d/%m/%Y")
        except Exception:
            return ""

    if isinstance(val, str):
        s = val.strip()
        if not s:
            return ""
        try:
            parsed = dateutil_parse(s, dayfirst=True)
            return parsed.strftime("%d/%m/%Y")
        except Exception:
            return ""

    return ""


def _workbook_invoice_number_cell(raw: object) -> str | None:
    if raw is None:
        return None
    try:
        if isinstance(raw, float) and pd.isna(raw):
            return None
    except Exception:
        pass
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = _norm_inv(str(raw))
    return s or None


def _invoice_number_is_summary_row(inv: str) -> bool:
    low = inv.lower()
    return any(
        sub in low
        for sub in ("sample size", "total", "summary", "count")
    )


def _excel_sheet_to_records(workbook_bytes: bytes, sheet_name: str) -> list[dict]:
    """
    Read a sheet into records, auto-detecting the header row by searching for 'Invoice Number'.
    This handles the generated workpaper format where some sheets have a note row above headers.
    """
    df_raw = pd.read_excel(
        io.BytesIO(workbook_bytes), sheet_name=sheet_name, header=None, dtype=object
    )
    if df_raw is None or df_raw.empty:
        return []

    header_row_idx: int | None = None
    for i in range(min(len(df_raw.index), 30)):
        row_vals = df_raw.iloc[i].tolist()
        if any(str(v).strip().lower() == "invoice number" for v in row_vals if v is not None):
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    headers = [str(v).strip() if v is not None else "" for v in df_raw.iloc[header_row_idx].tolist()]
    df = df_raw.iloc[header_row_idx + 1 :].copy()
    df.columns = headers
    df = df.dropna(how="all")
    if df.empty:
        return []
    return df.to_dict(orient="records")


def _load_sample_items_from_workpaper(workbook_bytes: bytes) -> list[dict]:
    _gl_amount_text_labels = frozenset({"sample size", "total value", "count"})

    def looks_like_invoice_ref(inv: str) -> bool:
        if not inv:
            return False
        if _invoice_number_is_summary_row(inv):
            return False
        return bool(re.match(r"^[A-Za-z0-9]", inv))

    items: list[dict] = []
    for sheet in ["MUS Sample", "Target Testing"]:
        try:
            records = _excel_sheet_to_records(workbook_bytes, sheet)
        except Exception:
            records = []
        if len(records) > 3:
            records = records[:-3]
        for r in records:
            raw_inv = _row_get_ignore_case(r, "Invoice Number")
            inv = _workbook_invoice_number_cell(raw_inv)
            print(f"[SHEET DEBUG] sheet={sheet} inv={inv or ''} all_keys={list(r.keys())[:8]}")
            if not inv or not looks_like_invoice_ref(inv):
                continue
            print(
                f"[FILTER CHECK] inv='{inv}' lower='{inv.strip().lower()}' in_set={inv.strip().lower() in ('sample size', 'sample size:', 'summary', 'total', 'count')}"
            )
            if inv.strip().lower() in ("sample size", "sample size:", "summary", "total", "count"):
                continue
            amt_raw = _row_get_ignore_case(r, "Amount")
            if isinstance(amt_raw, str) and amt_raw.strip().lower() in _gl_amount_text_labels:
                continue
            gl_amt = _safe_float(amt_raw)
            if gl_amt is None:
                continue
            date_cell = _row_gl_date_raw(r)
            gl_date_display = _gl_date_display_dd_mm_yyyy(date_cell)
            print(f"[GL DATE] invoice={inv} raw={date_cell} display={gl_date_display}")
            cust_raw = _row_get_ignore_case(r, "Customer")
            cust = _norm_inv(cust_raw) if cust_raw is not None and not (
                isinstance(cust_raw, float) and pd.isna(cust_raw)
            ) else ""
            items.append(
                {
                    "source_sheet": sheet,
                    "invoice_number": inv,
                    "gl_amount_ex_gst": gl_amt,
                    "date": gl_date_display,
                    "customer": cust or None,
                }
            )

    # De-duplicate by invoice number
    out: list[dict] = []
    seen: set[str] = set()
    for it in items:
        k = it["invoice_number"]
        if k in seen:
            continue
        seen.add(k)
        out.append(it)
    return out


def _azure_document_intelligence_client():
    endpoint = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT")
    key = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY")
    if not endpoint or "your_endpoint_here" in endpoint:
        raise RuntimeError("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT is missing or still set to placeholder.")
    if not key or "your_key_here" in key:
        raise RuntimeError("AZURE_DOCUMENT_INTELLIGENCE_KEY is missing or still set to placeholder.")

    try:
        from azure.ai.documentintelligence import DocumentIntelligenceClient
        from azure.core.credentials import AzureKeyCredential
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "Azure Document Intelligence SDK not installed. Install: pip install azure-ai-documentintelligence"
        ) from e

    return DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))


def _begin_analyze_document(client, model_id: str, file_bytes: bytes):
    """Run Azure Document Intelligence analyze and return the AnalyzeResult."""
    poller = None
    try:
        poller = client.begin_analyze_document(model_id, file_bytes)
    except TypeError:
        try:
            from azure.ai.documentintelligence.models import AnalyzeDocumentRequest

            poller = client.begin_analyze_document(
                model_id, AnalyzeDocumentRequest(bytes_source=file_bytes)
            )
        except Exception:
            poller = client.begin_analyze_document(model_id, document=file_bytes)
    return poller.result()


def _extract_text_with_azure_di(client, file_bytes: bytes) -> str:
    """
    Extract text using Azure Document Intelligence prebuilt-read.
    """
    result = _begin_analyze_document(client, "prebuilt-read", file_bytes)

    chunks: list[str] = []
    content = getattr(result, "content", None)
    if content:
        return str(content)

    pages = getattr(result, "pages", None) or []
    for page in pages:
        lines = getattr(page, "lines", None) or []
        for line in lines:
            txt = getattr(line, "content", None)
            if txt:
                chunks.append(str(txt))
    return "\n".join(chunks).strip()


def _document_field_scalar(df: object) -> object | None:
    """Best-effort scalar from an Azure Document Intelligence DocumentField."""
    if df is None:
        return None
    t = str(getattr(df, "type", "") or "").lower()
    if "currency" in t:
        vc = getattr(df, "value_currency", None)
        if vc is not None and getattr(vc, "amount", None) is not None:
            return float(vc.amount)
        return None
    if t == "number":
        vn = getattr(df, "value_number", None)
        return float(vn) if vn is not None else None
    if t == "integer":
        vi = getattr(df, "value_integer", None)
        return float(vi) if vi is not None else None
    if t == "date":
        vd = getattr(df, "value_date", None)
        if vd is not None:
            return vd.isoformat() if hasattr(vd, "isoformat") else str(vd)
        return None
    if t == "string":
        return getattr(df, "value_string", None) or None
    if t == "address":
        va = getattr(df, "value_address", None)
        if va is not None:
            parts = [
                getattr(va, "street_address", None),
                getattr(va, "city", None),
                getattr(va, "state", None),
                getattr(va, "postal_code", None),
            ]
            joined = ", ".join(str(p) for p in parts if p)
            if joined:
                return joined
        return getattr(df, "content", None)
    if t == "object":
        return getattr(df, "content", None)
    return getattr(df, "content", None)


def _parse_prebuilt_invoice_analyze_result(analyze_result: object) -> dict:
    """
    Map prebuilt-invoice AnalyzeResult to the app's invoice dict shape.
    """
    out: dict = {
        "vendor_name": None,
        "invoice_number": None,
        "customer": None,
        "date": None,
        "amount_ex_gst": None,
        "total_tax": None,
        "amount_inc_gst": None,
        "line_items": [],
    }
    docs = getattr(analyze_result, "documents", None) or []
    if not docs:
        return out
    doc = docs[0]
    fields = getattr(doc, "fields", None) or {}

    def grab(key: str) -> object | None:
        return _document_field_scalar(fields.get(key))

    out["vendor_name"] = _norm_inv(grab("VendorName")) or None
    out["customer"] = _norm_inv(grab("CustomerName")) or None
    out["invoice_number"] = _norm_inv(grab("InvoiceId")) or None
    inv_date = grab("InvoiceDate")
    out["date"] = str(inv_date) if inv_date is not None else None

    sub = grab("SubTotal")
    tax = grab("TotalTax")
    # Total (inc GST) can appear under different field names depending on layout.
    total = grab("InvoiceTotal")
    due = grab("AmountDue")
    total_alt = grab("TotalAmount")
    grand_total = grab("GrandTotal")
    out["amount_due"] = float(due) if due is not None else None
    out["amount_ex_gst"] = float(sub) if sub is not None else None
    out["total_tax"] = float(tax) if tax is not None else None
    total_inc = None
    for cand in (total, due, total_alt, grand_total):
        if cand is not None:
            total_inc = cand
            break
    out["amount_inc_gst"] = float(total_inc) if total_inc is not None else None

    items_field = fields.get("Items")
    line_items: list[dict] = []
    arr = getattr(items_field, "value_array", None) if items_field is not None else None
    if arr:
        for row in arr:
            obj = getattr(row, "value_object", None) or {}
            desc = None
            amt = None
            if isinstance(obj, dict):
                desc = _norm_inv(_document_field_scalar(obj.get("Description"))) or None
                amt = _document_field_scalar(obj.get("Amount"))
                if amt is None:
                    amt = _document_field_scalar(obj.get("AmountIncludingTax"))
            else:
                desc = getattr(row, "content", None)
            try:
                amt_f = float(amt) if amt is not None else None
            except Exception:
                amt_f = None
            if desc or amt_f is not None:
                line_items.append({"description": desc, "amount": amt_f})
    out["line_items"] = line_items

    # Derive ex-GST from inc if SubTotal missing (common on some layouts).
    if out["amount_ex_gst"] is None and out["amount_inc_gst"] is not None and out["total_tax"] is not None:
        out["amount_ex_gst"] = round(float(out["amount_inc_gst"]) - float(out["total_tax"]), 2)
    elif out["amount_ex_gst"] is None and out["amount_inc_gst"] is not None:
        out["amount_ex_gst"] = round(float(out["amount_inc_gst"]) / 1.1, 2)

    return out


def _invoice_merge_openai_fallback(client, parsed: dict, ocr_text: str) -> dict:
    """Fill gaps in prebuilt-invoice output using OpenAI on the same OCR text."""
    t = (ocr_text or "").strip()
    if not t:
        return parsed
    need = (
        not parsed.get("invoice_number")
        or parsed.get("amount_inc_gst") is None
        or not parsed.get("customer")
    )
    if not need:
        return parsed
    fb = _extract_invoice_fields_openai(client, t)
    out = dict(parsed)
    if not out.get("invoice_number"):
        out["invoice_number"] = fb.get("invoice_number")
    if out.get("amount_inc_gst") is None:
        out["amount_inc_gst"] = fb.get("amount_inc_gst")
    if not out.get("customer"):
        out["customer"] = fb.get("customer")
    if not out.get("date"):
        out["date"] = fb.get("date") or fb.get("invoice_date")
    if not out.get("description"):
        out["description"] = fb.get("description") or fb.get("line_items_summary")
    if out.get("amount_ex_gst") is None and fb.get("amount_ex_gst") is not None:
        try:
            out["amount_ex_gst"] = round(float(fb["amount_ex_gst"]), 2)
        except Exception:
            pass
    if out.get("total_tax") is None and fb.get("gst_amount") is not None:
        try:
            out["total_tax"] = round(float(fb["gst_amount"]), 2)
        except Exception:
            pass
    if out.get("amount_ex_gst") is None:
        amt_inc = out.get("amount_inc_gst")
        tax = out.get("total_tax")
        if amt_inc is not None and tax is not None:
            try:
                out["amount_ex_gst"] = round(float(amt_inc) - float(tax), 2)
            except Exception:
                out["amount_ex_gst"] = (float(amt_inc) / 1.1) if amt_inc is not None else None
        elif amt_inc is not None:
            try:
                out["amount_ex_gst"] = round(float(amt_inc) / 1.1, 2)
            except Exception:
                out["amount_ex_gst"] = None
    return out


def _best_gst_inc_ex_pair(candidate_amounts: list[float]) -> tuple[float, float] | None:
    """
    If any two distinct positive amounts satisfy larger ≈ 1.1 × smaller (within 5% of 1.1),
    return (inc_abs, ex_abs). Chooses the pair with ratio closest to 1.1.
    """
    uniq = sorted({round(abs(x), 2) for x in candidate_amounts if x is not None and abs(x) > 1e-9})
    best: tuple[float, float] | None = None
    best_err: float | None = None
    tol = 1.1 * 0.05
    for i, small in enumerate(uniq):
        for large in uniq[i + 1 :]:
            if small <= 0:
                continue
            err = abs(large / small - 1.1)
            if err <= tol:
                if best_err is None or err < best_err:
                    best_err = err
                    best = (large, small)
    return best


def _finalize_invoice_amounts(parsed: dict) -> dict:
    """
    Prefer explicit GST pairs: if two amounts match ~1.1×, larger is inc-GST and smaller ex-GST.
    Otherwise inc-GST is the largest amount among invoice fields and line aggregates; ex-GST
    prefers labelled SubTotal when strictly below that total; otherwise inc÷1.1.
    Swap if inc < ex; set gst_check_required when inc/ex deviates from 1.1 by more than 5%.
    """
    out = dict(parsed)
    inv_num_for_log = _norm_inv(out.get("invoice_number") or "") or None

    total_raw = _safe_float(out.get("amount_inc_gst"))
    sub_raw = _safe_float(out.get("amount_ex_gst"))
    due_raw = _safe_float(out.get("amount_due"))

    sign = -1 if (total_raw is not None and total_raw < 0) or (sub_raw is not None and sub_raw < 0) else 1

    line_amts: list[float] = []
    for row in out.get("line_items") or []:
        if not isinstance(row, dict):
            continue
        v = _safe_float(row.get("amount"))
        if v is not None and abs(v) > 0:
            line_amts.append(abs(v))
    line_sum = round(sum(line_amts), 2) if line_amts else None
    line_max = max(line_amts) if line_amts else None

    total_abs = abs(total_raw) if total_raw is not None else None
    sub_abs = abs(sub_raw) if sub_raw is not None else None
    due_abs = abs(due_raw) if due_raw is not None else None

    gst_candidates: list[float] = []
    for x in (total_abs, due_abs, sub_abs, line_sum, line_max):
        if x is not None:
            gst_candidates.append(float(x))
    gst_candidates.extend(line_amts)
    amounts_found = sorted({round(abs(x), 2) for x in gst_candidates if x is not None and abs(x) > 1e-9})

    pair = _best_gst_inc_ex_pair(gst_candidates)
    inc_abs: float | None = None
    ex_abs: float | None = None
    if pair:
        inc_abs, ex_abs = pair
    else:
        pool = [x for x in [total_abs, due_abs, sub_abs, line_sum, line_max] if x is not None]
        inc_abs = max(pool) if pool else None

        if inc_abs is None:
            pass
        elif sub_abs is not None and sub_abs < inc_abs - 1e-9:
            ex_abs = sub_abs
        else:
            ex_abs = round(inc_abs / 1.1, 2)

        if inc_abs is None and ex_abs is None and sub_abs is not None:
            ex_abs = sub_abs
            inc_abs = round(ex_abs * 1.1, 2)

    swapped_basic = False
    if inc_abs is not None and ex_abs is not None and inc_abs < ex_abs:
        inc_abs, ex_abs = ex_abs, inc_abs
        swapped_basic = True

    # Preserve labelled/existing ex-GST if it appears elsewhere in the amounts pool.
    # If any other amount is ~ (inc/1.1) within 2%, use it rather than recomputing.
    if inc_abs is not None and inc_abs > 0:
        target_ex = inc_abs / 1.1
        best_ex: float | None = None
        best_err: float | None = None
        for cand in amounts_found:
            if cand <= 0:
                continue
            if abs(cand - inc_abs) < 0.01:
                continue
            err = abs(cand - target_ex) / max(target_ex, 1e-9)
            if err <= 0.02:
                if best_err is None or err < best_err:
                    best_err = err
                    best_ex = cand
        if best_ex is not None:
            ex_abs = best_ex

    swapped_ratio = False
    ratio_disp: float | None = None
    if inc_abs is not None and ex_abs is not None and ex_abs > 0 and inc_abs > 0:
        ratio_disp = inc_abs / ex_abs
        if not (1.05 <= ratio_disp <= 1.15):
            ratio_if_swapped = ex_abs / inc_abs
            if 1.05 <= ratio_if_swapped <= 1.15:
                inc_abs, ex_abs = ex_abs, inc_abs
                swapped_ratio = True
                ratio_disp = inc_abs / ex_abs
        print(f"[AMOUNT CHECK] inc={inc_abs} ex={ex_abs} ratio={ratio_disp} swapped={swapped_ratio}")

    swapped_any = swapped_basic or swapped_ratio
    print(
        f"[AMOUNT DEBUG] inv={inv_num_for_log} all_amounts_found={amounts_found} "
        f"selected_inc={inc_abs} selected_ex={ex_abs} ratio={ratio_disp} swapped={swapped_any}"
    )

    if inc_abs is not None:
        out["amount_inc_gst"] = round(inc_abs * sign, 2)
    if ex_abs is not None:
        out["amount_ex_gst"] = round(ex_abs * sign, 2)

    tax_keep = _safe_float(out.get("total_tax"))
    if inc_abs is not None and ex_abs is not None:
        derived_tax = round(inc_abs - ex_abs, 2) * sign
        if tax_keep is None or abs(abs(float(tax_keep)) - abs(derived_tax)) > inc_abs * 0.02:
            out["total_tax"] = derived_tax

    gst_check = False
    ai = abs(float(out["amount_inc_gst"])) if out.get("amount_inc_gst") is not None else None
    ae = abs(float(out["amount_ex_gst"])) if out.get("amount_ex_gst") is not None else None
    if ai is not None and ae is not None and ae > 0.01:
        ratio = ai / ae
        if abs(ratio - 1.1) > 1.1 * 0.05:
            gst_check = True
    out["gst_check_required"] = gst_check

    return out


def _normalized_money_amount_for_bank(val: object) -> float | None:
    """Strip commas, currency codes/symbols, spaces; return absolute scalar for comparison."""
    if val is None:
        return None
    try:
        if isinstance(val, float) and pd.isna(val):
            return None
    except Exception:
        pass
    if isinstance(val, (int, float)):
        try:
            return abs(float(val))
        except Exception:
            return None
    if isinstance(val, str):
        t = val.strip().upper().replace(",", "").replace(" ", "").replace("\u00a0", "")
        for sym in ("AUD", "USD", "NZD", "EUR", "GBP", "CAD"):
            t = t.replace(sym, "")
        for ch in "$€£¥":
            t = t.replace(ch, "")
        if not t:
            return None
        if re.search(r"[A-Za-z]", t):
            return None
        try:
            return abs(float(t))
        except Exception:
            return None
    try:
        return abs(float(val))
    except Exception:
        return None


def _bank_amounts_align(inv_abs: float, tx_abs: float, *, relative_tol: float = 0.01) -> bool:
    """Within relative_tol (default 1%) or trivial absolute difference after normalization."""
    if inv_abs <= 0 or tx_abs <= 0:
        return False
    m = max(inv_abs, tx_abs, 1e-9)
    return abs(inv_abs - tx_abs) / m <= relative_tol or abs(inv_abs - tx_abs) <= 0.05


def _bank_tx_dates_within_or_missing(inv_date_val: object, tx_date_val: object, max_days: int) -> bool:
    if not inv_date_val or not tx_date_val:
        return True
    try:
        d1 = pd.to_datetime(inv_date_val, dayfirst=True).date()
        d2 = pd.to_datetime(tx_date_val, dayfirst=True).date()
        return abs((d2 - d1).days) <= max_days
    except Exception:
        return True


def _bank_tx_positive_credit_value(tx: dict) -> float | None:
    """
    Choose a positive receipt value from a bank transaction:
    - If credit_amount exists and > 0, use it
    - Else if amount exists and > 0, use it
    - If only debit_amount exists (or values are <=0), return None
    """
    if not tx:
        return None
    c = _safe_float(tx.get("credit_amount"))
    if c is not None and c > 0:
        return float(c)
    a = _safe_float(tx.get("amount"))
    if a is not None and a > 0:
        return float(a)
    # Explicit debit-only rows should not match receipts.
    d = _safe_float(tx.get("debit_amount"))
    if d is not None and d > 0:
        return None
    return None


def _normalize_for_bank_fuzzy_compare(text: str) -> str:
    """Lowercase, strip common company suffixes, replace punctuation with spaces."""
    s = _norm_inv(text).lower()
    for pat in (
        r"\bpty\s+ltd\b",
        r"\blimited\b",
        r"\bltd\b",
        r"\bgroup\b",
        r"\bincorporated\b",
        r"\binc\b",
        r"\bco\b",
        r"\bpty\b",
    ):
        s = re.sub(pat, " ", s, flags=re.I)
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _partial_customer_word_five_plus(customer_name: str | None, bank_description: str | None) -> bool:
    """Any normalized customer token of length ≥5 appears as a substring of the bank description."""
    if not customer_name or not bank_description:
        return False
    nc = _normalize_for_bank_fuzzy_compare(customer_name)
    bd = bank_description.lower()
    return any(len(tok) >= 5 and tok in bd for tok in nc.split())


def _fuzzy_customer_match(customer_name: str | None, bank_description: str | None) -> bool:
    """
    Token overlap on 4+ char tokens: True if ≥30% of customer tokens appear in bank tokens, or any
    customer token of length ≥6 appears in bank tokens / collapsed bank text.
    """
    if not customer_name or not bank_description:
        return False
    nc = _normalize_for_bank_fuzzy_compare(customer_name)
    nb = _normalize_for_bank_fuzzy_compare(bank_description)
    if not nc or not nb:
        return False
    cust_tokens = [t for t in nc.split() if len(t) >= 4]
    if not cust_tokens:
        return False
    bank_tokens = {t for t in nb.split() if len(t) >= 4}
    bank_collapsed = nb.replace(" ", "")
    overlap = sum(1 for t in cust_tokens if t in bank_tokens)
    need_overlap = max(1, math.ceil(len(cust_tokens) * 0.3))
    overlap_ok = overlap >= need_overlap
    long_token_hit = any(
        len(t) >= 6 and (t in bank_tokens or t in bank_collapsed) for t in cust_tokens
    )
    return overlap_ok or long_token_hit


def _customer_matches_bank_description(customer_name: str | None, bank_description: str | None) -> bool:
    """Fuzzy overlap / long-token rules OR any 5+ character customer word found in the bank text."""
    if not customer_name or not bank_description:
        return False
    if _partial_customer_word_five_plus(customer_name, bank_description):
        return True
    return _fuzzy_customer_match(customer_name, bank_description)


def _normalized_customer_tokens_for_log(customer_name: str | None) -> list[str]:
    if not customer_name:
        return []
    nc = _normalize_for_bank_fuzzy_compare(customer_name)
    return [t for t in nc.split() if len(t) >= 4]


def _invoice_number_in_bank_description(invoice_number: str | None, bank_description: str | None) -> bool:
    """True if the invoice reference appears in the bank narrative (spacing/punctuation tolerant)."""
    if not invoice_number or not bank_description:
        return False
    inv = _norm_inv(invoice_number).strip()
    if not inv:
        return False
    desc = bank_description.lower()
    inv_l = inv.lower()
    if inv_l in desc:
        return True
    inv_compact = re.sub(r"[^a-z0-9]+", "", inv_l)
    desc_compact = re.sub(r"[^a-z0-9]+", "", desc)
    return bool(inv_compact) and inv_compact in desc_compact


def _normalize_invoice_date_from_extract(s: str) -> str:
    """
    Normalize model-returned invoice_date to DD/MM/YYYY.
    - YYYY-MM-DD -> DD/MM/YYYY
    - Slash dates: if second segment > 12 it cannot be month under DD/MM — treat as MM/DD and swap.
      If first segment > 12 it cannot be month under MM/DD — treat as DD/MM (no swap).
      If both segments <= 12, keep order (ambiguous; assumed DD/MM).
    """
    s = (s or "").strip()
    if not s:
        return ""

    iso_m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)
    if iso_m:
        try:
            y, mo, d = int(iso_m.group(1)), int(iso_m.group(2)), int(iso_m.group(3))
            return dt.date(y, mo, d).strftime("%d/%m/%Y")
        except Exception:
            return s

    parts = [p.strip() for p in s.split("/")]
    if len(parts) != 3:
        return s
    try:
        a, b, y = int(parts[0]), int(parts[1]), int(parts[2])
    except ValueError:
        return s

    if y < 100:
        y += 2000 if y < 70 else 1900

    if a > 12:
        day, month = a, b
    elif b > 12:
        month, day = a, b
    else:
        day, month = a, b

    try:
        return dt.date(y, month, day).strftime("%d/%m/%Y")
    except Exception:
        return s


def _openai_client():
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or "your_openai_key_here" in api_key:
        raise RuntimeError("OPENAI_API_KEY is missing or still set to placeholder.")
    try:
        from openai import OpenAI
    except Exception as e:  # pragma: no cover
        raise RuntimeError("OpenAI SDK not installed. Install: pip install openai") from e
    return OpenAI(api_key=api_key)


def _extract_invoice_fields_openai(client, invoice_text: str) -> dict:
    invoice_text = (invoice_text or "").strip()
    empty = {
        "invoice_number": None,
        "invoice_date": None,
        "customer": None,
        "amount_inc_gst": None,
        "amount_ex_gst": None,
        "gst_amount": None,
        "line_items_summary": None,
        "date": None,
        "description": None,
    }
    if not invoice_text:
        return dict(empty)

    system = (
        "You are an expert invoice data extractor for Australian audit firms. "
        "You extract structured data from invoices regardless of layout or format. "
        "Always return valid JSON only, no other text."
    )

    user = f"""Extract data from this invoice. Follow these rules strictly:
AMOUNT RULES:
- Find every dollar amount in the document
- amount_inc_gst = the LARGEST amount (this is always the total including GST)
- The invoice total may be labelled: "Amount Due" (including GST), "Invoice Total", "Total", "Amount Payable", "Total Due"
- amount_ex_gst = amount_inc_gst divided by 1.1 (unless clearly labelled)
- gst_amount = amount_inc_gst minus amount_ex_gst
- For construction/progress claims: use "This claim" or "Amount this claim" or "Claimed this period" as amount_inc_gst
- NEVER return the smallest amount as amount_inc_gst
DATE RULES:
- Find any date labelled: Invoice Date, Date, Tax Date, Bill Date, Service Date, Issue Date, Tax Invoice Date, Period
- Return the invoice issue date in DD/MM/YYYY format
- Convert MM/DD/YYYY to DD/MM/YYYY if needed
INVOICE NUMBER RULES:
- Look for: Invoice Number, Invoice No, Invoice #, Reference No, Ref No, Claim No, Tax Invoice No
- Return only the alphanumeric code
CUSTOMER RULES:
- Look for: Bill To, Customer, Client, Attention, To, Sold To
- Customer may be at top OR bottom of page
Return ONLY this JSON structure with no other text:
{{"invoice_number": "...", "invoice_date": "DD/MM/YYYY", "customer": "...", "amount_inc_gst": 0.00, "amount_ex_gst": 0.00, "gst_amount": 0.00, "line_items_summary": "..."}}
Invoice text to extract from:
{invoice_text[:20000]}
Return null for any field you cannot find with confidence."""
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        response_format={"type": "json_object"},
    )
    raw = (resp.choices[0].message.content or "").strip()
    data = json.loads(raw) if raw else {}
    inv_date_raw = data.get("invoice_date")
    if inv_date_raw is not None and str(inv_date_raw).strip():
        normalized_date = _normalize_invoice_date_from_extract(str(inv_date_raw))
        if normalized_date:
            data["invoice_date"] = normalized_date
    print(
        f"[EXTRACTION] Invoice: {data.get('invoice_number')} | Inc GST: {data.get('amount_inc_gst')} | "
        f"Ex GST: {data.get('amount_ex_gst')} | Date: {data.get('invoice_date')} | Customer: {data.get('customer')}"
    )

    def _coerce_float(key: str) -> float | None:
        v = data.get(key)
        if v is None:
            return None
        try:
            return float(v)
        except Exception:
            return None

    inv = _norm_inv(data.get("invoice_number"))
    inv = inv or None
    invoice_date = data.get("invoice_date")
    invoice_date_s = _norm_inv(str(invoice_date)) if invoice_date not in (None, "") else None
    customer = _norm_inv(data.get("customer")) or None
    line_summary = _norm_inv(data.get("line_items_summary")) or None
    amt_inc = _coerce_float("amount_inc_gst")
    amt_ex = _coerce_float("amount_ex_gst")
    gst_amt = _coerce_float("gst_amount")

    return {
        "invoice_number": inv,
        "invoice_date": invoice_date_s,
        "customer": customer,
        "amount_inc_gst": amt_inc,
        "amount_ex_gst": amt_ex,
        "gst_amount": gst_amt,
        "line_items_summary": line_summary,
        "date": invoice_date_s,
        "description": line_summary,
    }


def _extract_remittance_openai(client, text: str) -> dict:
    text = (text or "").strip()
    if not text:
        return {
            "customer_name": None,
            "payment_date": None,
            "invoice_references": [],
            "amounts_paid": [],
            "total_amount": None,
        }

    system = (
        "You extract remittance / payment advice details from OCR text. "
        "Return ONLY valid JSON. Use null for unknown scalars, empty arrays if absent."
    )
    user = (
        "Extract:\n"
        "- customer_name (string)\n"
        "- payment_date (YYYY-MM-DD if possible)\n"
        "- invoice_references (array of invoice numbers as strings)\n"
        "- amounts_paid (array of numbers, same length as invoice_references; "
        "amount paid per referenced invoice in order)\n"
        "- total_amount (number)\n\n"
        f"TEXT:\n{text[:25000]}"
    )
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        response_format={"type": "json_object"},
    )
    raw = (resp.choices[0].message.content or "").strip()
    data = json.loads(raw) if raw else {}

    refs = data.get("invoice_references") or []
    amts = data.get("amounts_paid") or []
    if not isinstance(refs, list):
        refs = []
    if not isinstance(amts, list):
        amts = []
    refs = [_norm_inv(x) for x in refs if x is not None and str(x).strip()]
    nums: list[float | None] = []
    for i in range(len(refs)):
        v = amts[i] if i < len(amts) else None
        try:
            nums.append(float(v) if v is not None else None)
        except Exception:
            nums.append(None)

    total_amt = data.get("total_amount")
    try:
        total_amt = float(total_amt) if total_amt is not None else None
    except Exception:
        total_amt = None

    return {
        "customer_name": _norm_inv(data.get("customer_name")) or None,
        "payment_date": _norm_inv(data.get("payment_date")) or None,
        "invoice_references": refs,
        "amounts_paid": nums,
        "total_amount": total_amt,
    }


def _extract_bank_transactions_openai(client, bank_text: str) -> list[dict]:
    bank_text = (bank_text or "").strip()
    if not bank_text:
        return []

    system = (
        "You extract bank statement transaction rows from OCR text. Return ONLY valid JSON.\n"
        'Each row must have exactly these keys: '
        '"date" (YYYY-MM-DD), "description" (string), '
        '"debit_amount" (number or null), "credit_amount" (number or null), "balance" (number or null).\n'
        "Use null when a column does not apply or is unreadable. "
        "For receipts/deposits set credit_amount; for payments out set debit_amount. "
        'Output: {"transactions": [ {...}, ... ]}'
    )
    user = f"BANK STATEMENT TEXT:\n{bank_text[:25000]}"
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        response_format={"type": "json_object"},
    )
    raw = (resp.choices[0].message.content or "").strip()
    data = json.loads(raw) if raw else {}
    txs = data.get("transactions") or []
    out: list[dict] = []
    for t in txs:
        if not isinstance(t, dict):
            continue
        debit = _safe_float(t.get("debit_amount"))
        credit = _safe_float(t.get("credit_amount"))
        bal = _safe_float(t.get("balance"))
        legacy_amt = _safe_float(t.get("amount"))
        # Receipt amount for matching inbound payments (inc GST invoice totals).
        receipt_amt: float | None = None
        if credit is not None and float(credit) > 0:
            receipt_amt = float(credit)
        elif legacy_amt is not None and legacy_amt > 0:
            receipt_amt = legacy_amt
        out.append(
            {
                "date": _norm_inv(t.get("date")) or None,
                "description": _norm_inv(t.get("description")) or None,
                "debit_amount": debit,
                "credit_amount": credit,
                "balance": bal,
                "amount": receipt_amt,
            }
        )
    return out


def _line_items_summary(invoice: dict | None, max_len: int = 32000) -> str:
    if not invoice:
        return ""
    items = invoice.get("line_items") or []
    parts: list[str] = []
    for row in items:
        if not isinstance(row, dict):
            continue
        d = _norm_inv(row.get("description"))
        amt = row.get("amount")
        if amt is not None:
            try:
                parts.append(f"{d}: {float(amt):,.2f}".strip(": "))
            except Exception:
                parts.append(f"{d}: {amt}".strip(": "))
        elif d:
            parts.append(d)
    s = " | ".join(parts)
    return s[:max_len] if len(s) > max_len else s


def _cell_date_gl(val: object):
    """GL column display value as DD/MM/YYYY string (workpaper + API results)."""
    return _gl_date_display_dd_mm_yyyy(val)


def _cell_date_iso(val: object):
    if val is None or val == "":
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def _choose_best_invoice(sample_item: dict, invoices: list[dict]) -> dict | None:
    inv_key = _norm_inv(sample_item.get("invoice_number")).lower()
    if not inv_key:
        return None
    candidates = [x for x in invoices if _norm_inv(x.get("invoice_number")).lower() == inv_key]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    gl_amt = sample_item.get("gl_amount_ex_gst")
    if gl_amt is None:
        return candidates[0]

    def score(x: dict) -> float:
        ex = x.get("amount_ex_gst")
        if ex is None:
            return float("inf")
        return abs(float(ex) - float(gl_amt))

    return sorted(candidates, key=score)[0]


def _invoice_ref_keys(inv: str) -> set[str]:
    """Normalization variants for comparing invoice numbers across sources."""
    n = _norm_inv(inv).strip()
    if not n:
        return set()
    low = n.lower()
    return {low, re.sub(r"\s+", "", low)}


def _remittance_covers_invoice(remittances: list[dict], inv_label: str) -> tuple[bool, bool, str | None]:
    """
    Returns (listed_on_remittance, amount_aligned_with_invoice_inc, payment_date_from_remittance).
    amount_aligned is True if any line pairs this invoice with a paid amount matching invoice inc GST
    (checked by caller) — here we only signal listing + optional per-line amount match vs caller totals.
    """
    target_keys = _invoice_ref_keys(inv_label)
    if not target_keys:
        return (False, False, None)

    listed = False
    amount_ok = False
    pay_date: str | None = None
    for rm in remittances:
        refs = rm.get("invoice_references") or []
        amts = rm.get("amounts_paid") or []
        pd = _norm_inv(rm.get("payment_date")) or None
        for i, ref in enumerate(refs):
            rk = _invoice_ref_keys(ref)
            if not rk & target_keys:
                continue
            listed = True
            pay_date = pd or pay_date
            if i < len(amts) and amts[i] is not None:
                amount_ok = True
        if pay_date is None:
            pay_date = _norm_inv(rm.get("payment_date")) or None

    return (listed, amount_ok, pay_date)


def _bank_tx_credit_display(tx: dict | None) -> float | None:
    if not tx:
        return None
    return _bank_tx_positive_credit_value(tx)


def _bank_match_for_sample(
    sample_item: dict,
    invoice: dict | None,
    bank_txs: list[dict],
    remittances: list[dict],
) -> tuple[str, str | None, str, str | None, str | None, float | None]:
    """
    Returns (bank_match, bank_receipt_date, confidence, document_type,
             bank_description, bank_credit_amount).

    document_type is one of:
        remittance_match — remittance lines list this invoice and/or remittance waives the customer-name
            check on a bank line for that invoice; or remittance-only proof when listed.
        invoice_direct — normalized amounts align (≤1%); fuzzy/partial customer match; date ≤45d when dates known.
        bank_statement_match — invoice number appears in the bank description and amounts/date gate pass.
        amount_date_fallback — amount aligns within 1% and tx date within 60d of invoice date when dates known;
            used when customer name and invoice ref do not match.
    """
    if invoice is None:
        return ("No", None, "Low", None, None, None)

    amount_inc = invoice.get("amount_inc_gst")
    inv_date = invoice.get("date")
    customer = _norm_inv(invoice.get("customer") or sample_item.get("customer") or "")
    inv_num = _norm_inv(sample_item.get("invoice_number") or invoice.get("invoice_number") or "")

    if amount_inc is None:
        return ("No", None, "Low", None, None, None)

    amount_inc_f = float(amount_inc)
    na_inv = _normalized_money_amount_for_bank(amount_inc_f)
    if na_inv is None or na_inv <= 0:
        return ("No", None, "Low", None, None, None)

    listed, _amt_pair_present, remit_pay_date = _remittance_covers_invoice(remittances, inv_num)

    # Remittance lines with explicit amounts vs invoice total (primary remittance proof).
    for rm in remittances:
        refs = rm.get("invoice_references") or []
        amts = rm.get("amounts_paid") or []
        pd = _norm_inv(rm.get("payment_date")) or None
        tk = _invoice_ref_keys(inv_num)
        for i, ref in enumerate(refs):
            if not (_invoice_ref_keys(ref) & tk):
                continue
            line_amt = amts[i] if i < len(amts) else None
            if line_amt is None:
                continue
            na_line = _normalized_money_amount_for_bank(line_amt)
            if na_line is not None and _bank_amounts_align(na_line, na_inv):
                return ("Yes", pd or remit_pay_date, "High", "remittance_match", None, None)

    cust_tokens_log = _normalized_customer_tokens_for_log(customer)
    print(
        f"[bank-match] sample_invoice={inv_num!r} customer={customer!r} "
        f"norm_customer_tokens_4plus={cust_tokens_log!r} inc_gst_normalized={na_inv!r}"
    )

    matches_meta: list[tuple[dict, str]] = []

    for tx in bank_txs:
        desc_raw = _norm_inv(tx.get("description"))
        tx_date = tx.get("date")
        invoice_ref_ok = _invoice_number_in_bank_description(inv_num, desc_raw)
        date_ok_45 = _bank_tx_dates_within_or_missing(inv_date, tx_date, 45)
        date_ok_60 = _bank_tx_dates_within_or_missing(inv_date, tx_date, 60)

        # If invoice reference appears, allow match even when amount is missing/unconfirmed.
        # Note: "Invoice reference found, amount unconfirmed."
        if invoice_ref_ok and date_ok_60:
            matches_meta.append((tx, "invoice_ref_unconfirmed_amount"))
            print(
                f"  bank_line tx_date={tx_date!r} desc={desc_raw[:180]!r} matched=True "
                f"reasons={['invoice_ref_in_desc', 'amount_unconfirmed', 'classified=invoice_ref_unconfirmed_amount']!r}"
            )
            continue

        tx_pos = _bank_tx_positive_credit_value(tx)
        na_tx = _normalized_money_amount_for_bank(tx_pos) if tx_pos is not None else None
        if na_tx is None or na_tx <= 0:
            print(f"  skip_no_bank_amount tx_date={tx_date!r} desc_snip={desc_raw[:140]!r}")
            continue

        if not _bank_amounts_align(na_inv, na_tx):
            print(
                f"  skip_amount_align_fail bank_amt_norm={na_tx!r} inv_norm={na_inv!r} "
                f"tx_date={tx_date!r} desc_snip={desc_raw[:140]!r}"
            )
            continue

        customer_ok = _customer_matches_bank_description(customer, desc_raw)
        remit_ok = listed

        amount_only_ok = (
            not remit_ok and not invoice_ref_ok and not customer_ok and date_ok_60
        )

        doc_type: str | None = None
        if remit_ok and date_ok_45:
            doc_type = "remittance_match"
        elif invoice_ref_ok and date_ok_45:
            doc_type = "bank_statement_match"
        elif customer_ok and date_ok_45:
            doc_type = "invoice_direct"
        elif amount_only_ok:
            doc_type = "amount_date_fallback"

        reasons: list[str] = []
        if customer_ok:
            reasons.append("customer_text_match")
        if invoice_ref_ok:
            reasons.append("invoice_ref_in_desc")
        if remit_ok:
            reasons.append("remittance_list_flag")
        if amount_only_ok:
            reasons.append("amount_date_fallback_60d")
        if doc_type:
            reasons.append(f"classified={doc_type}")
        else:
            if remit_ok and not date_ok_45:
                reasons.append("rejected_date_window_45d(remittance_path)")
            elif invoice_ref_ok and not date_ok_45:
                reasons.append("rejected_date_window_45d(invoice_ref)")
            elif customer_ok and not date_ok_45:
                reasons.append("rejected_date_window_45d(customer)")
            else:
                reasons.append("no_match_rule")

        matched_here = doc_type is not None
        print(
            f"  bank_line bank_amt_norm={na_tx!r} tx_date={tx_date!r} "
            f"desc={desc_raw[:180]!r} matched={matched_here} reasons={reasons!r}"
        )

        if doc_type:
            matches_meta.append((tx, doc_type))

    if not matches_meta:
        # Listed on remittance but bank lines unclear — still treat as matched per audit trail.
        if listed:
            print(f"[bank-match] sample_invoice={inv_num!r} outcome=Yes confidence=Medium (remittance_only)")
            return ("Yes", remit_pay_date, "Medium", "remittance_match", None, None)
        print(f"[bank-match] sample_invoice={inv_num!r} outcome=No confidence=Low (no_bank_line_match)")
        return ("No", None, "Low", None, None, None)

    # Prefer evidence: remittance_match > invoice_direct > bank_statement_match > amount_date_fallback
    priority = {
        "remittance_match": 0,
        "invoice_direct": 1,
        "bank_statement_match": 2,
        "invoice_ref_unconfirmed_amount": 3,
        "amount_date_fallback": 4,
    }

    def sort_key(item: tuple[dict, str]) -> tuple[int, str]:
        tx, dt = item
        return (priority.get(dt, 9), str(tx.get("date") or ""))

    matches_meta.sort(key=sort_key)
    best_tx, best_dt = matches_meta[0]

    conf = "High"
    if len(matches_meta) > 1 or best_dt in {"amount_date_fallback", "invoice_ref_unconfirmed_amount"}:
        conf = "Medium"
    outcome = "Yes" if len(matches_meta) == 1 else "Partial"
    print(
        f"[bank-match] sample_invoice={inv_num!r} outcome={outcome} confidence={conf} "
        f"best_document_type={best_dt!r} candidate_lines={len(matches_meta)}"
    )

    b_desc = _norm_inv(best_tx.get("description")) or None
    b_cred = _bank_tx_credit_display(best_tx)

    return (outcome, best_tx.get("date"), conf, best_dt, b_desc, b_cred)


async def _async_process_uploaded_pdf(
    di_client: object,
    oai: object,
    filename: str,
    file_bytes: bytes,
) -> dict | None:
    """OCR + extraction for one invoice PDF (Azure prebuilt-invoice + optional OpenAI gap-fill)."""
    if not file_bytes:
        return None
    fn = filename or ""
    inv_an = await asyncio.to_thread(_begin_analyze_document, di_client, "prebuilt-invoice", file_bytes)
    parsed = _parse_prebuilt_invoice_analyze_result(inv_an)
    ocr_text = getattr(inv_an, "content", None) or ""
    if not str(ocr_text).strip():
        ocr_text = await asyncio.to_thread(_extract_text_with_azure_di, di_client, file_bytes)
    parsed = await asyncio.to_thread(_invoice_merge_openai_fallback, oai, parsed, str(ocr_text))
    parsed = _finalize_invoice_amounts(parsed)
    parsed["source_filename"] = fn
    return parsed


async def _async_process_remittance_pdf(
    di_client: object,
    oai: object,
    filename: str,
    file_bytes: bytes,
) -> dict | None:
    """Remittance advice PDFs: Azure Read OCR + OpenAI structured extraction."""
    if not file_bytes:
        return None
    fn = filename or ""
    text = await asyncio.to_thread(_extract_text_with_azure_di, di_client, file_bytes)
    rm = await asyncio.to_thread(_extract_remittance_openai, oai, text)
    rm["source_filename"] = fn
    return rm


async def _async_process_bank_statement(di_client: object, oai: object, bank_bytes: bytes) -> list[dict]:
    bank_text = await asyncio.to_thread(_extract_text_with_azure_di, di_client, bank_bytes)
    return await asyncio.to_thread(_extract_bank_transactions_openai, oai, bank_text)


async def _async_merge_bank_transactions(di_client: object, oai: object, bank_reads: list[bytes]) -> list[dict]:
    """OCR + extract each bank PDF in parallel; concatenate transactions into one pool."""
    chunks = [b for b in bank_reads if b]
    if not chunks:
        return []
    parts = await asyncio.gather(
        *[_async_process_bank_statement(di_client, oai, b) for b in chunks]
    )
    merged: list[dict] = []
    for lst in parts:
        merged.extend(lst)
    return merged


def _invoice_gst_component(best_inv: dict | None, ex_gst: float | None, inc_gst: float | None) -> float | None:
    """GST amount: prefer TotalTax from invoice model; else inc − ex."""
    if best_inv is not None:
        tt = best_inv.get("total_tax")
        if tt is not None:
            try:
                return round(float(tt), 2)
            except Exception:
                pass
    if inc_gst is not None and ex_gst is not None:
        try:
            return round(float(inc_gst) - float(ex_gst), 2)
        except Exception:
            pass
    return None


def _bank_receipt_vs_invoice_label(
    bank_match: str,
    bank_cred: float | None,
    invoice_inc: float | None,
) -> str:
    """Compare matched bank credit to invoice total (inc GST)."""
    if bank_match == "No" or invoice_inc is None:
        return "No Match"
    if bank_cred is None:
        return "No Match"
    try:
        bc = float(bank_cred)
        inv = float(invoice_inc)
    except Exception:
        return "No Match"
    if abs(bc - inv) <= 0.01:
        return "Exact"
    if bc > inv + 0.01:
        return "Over"
    return "Under"


def _add_testing_results_sheet(workpaper_bytes: bytes, results: list[dict]) -> bytes:
    """
    Add/replace a 'Testing Results' sheet in the provided workpaper bytes.
    Styled to match the existing navy header workpaper style.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workpaper_bytes))

    sheet_name = "Testing Results"
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Invoice Number (GL)",
        "GL Amount (ex GST)",
        "GL Date",
        "GL Customer",
        "Invoice Number Extracted",
        "Invoice Date Extracted",
        "Invoice Customer Extracted",
        "Invoice Amount ex GST (SubTotal / OCR / inc÷1.1)",
        "Invoice GST Amount (TotalTax or inc − ex)",
        "Invoice Total (inc GST)",
        "GST check required (inc÷ex ≈ 1.1 ±5%)",
        "Amount Variance (GL vs Invoice ex GST)",
        "Line Items Summary",
        "Bank Match",
        "Bank Receipt Date",
        "Bank Description",
        "Bank Credit Amount",
        "Bank Receipt vs Invoice",
        "Document Type",
        "Match Confidence",
        "Performance Obligation",
        "Overall Result (Pass/Fail/Exception)",
        "Auditor Notes",
    ]
    _write_header(ws, 1, headers)

    vs_fill = {
        "Exact": PatternFill("solid", fgColor="E8F5E9"),
        "Over": PatternFill("solid", fgColor="FFF8E1"),
        "Under": PatternFill("solid", fgColor="FFEBEE"),
        "No Match": PatternFill("solid", fgColor="F5F5F5"),
    }

    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r.get("invoice_number"))
        ws.cell(row=i, column=2, value=r.get("gl_amount_ex_gst"))
        ws.cell(row=i, column=3, value=r.get("gl_date"))
        ws.cell(row=i, column=4, value=r.get("gl_customer"))
        ws.cell(row=i, column=5, value=r.get("invoice_number_extracted"))
        ws.cell(row=i, column=6, value=r.get("invoice_date_extracted"))
        ws.cell(row=i, column=7, value=r.get("invoice_customer_extracted"))
        ws.cell(row=i, column=8, value=r.get("invoice_amount_ex_gst"))
        ws.cell(row=i, column=9, value=r.get("invoice_gst_amount"))
        ws.cell(row=i, column=10, value=r.get("invoice_total_inc_gst"))
        ws.cell(
            row=i,
            column=11,
            value=("Yes" if r.get("gst_check_required") else ""),
        )
        ws.cell(row=i, column=12, value=r.get("amount_variance"))
        ws.cell(row=i, column=13, value=r.get("line_items_summary"))
        ws.cell(row=i, column=14, value=r.get("bank_match"))
        ws.cell(row=i, column=15, value=r.get("bank_receipt_date"))
        ws.cell(row=i, column=16, value=r.get("bank_description"))
        ws.cell(row=i, column=17, value=r.get("bank_credit_amount"))
        bvi = r.get("bank_receipt_vs_invoice")
        c18 = ws.cell(row=i, column=18, value=bvi)
        if bvi in vs_fill:
            c18.fill = vs_fill[bvi]
        ws.cell(row=i, column=19, value=r.get("document_type"))
        ws.cell(row=i, column=20, value=r.get("match_confidence"))
        ws.cell(row=i, column=21, value=r.get("performance_obligation"))
        ws.cell(row=i, column=22, value=r.get("overall_result"))
        ws.cell(row=i, column=23, value=r.get("auditor_notes"))

    max_col = len(headers)
    _apply_table_style(ws, header_row=1, min_col=1, max_col=max_col)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

    amt_cols = {2, 8, 9, 10, 12, 17}
    date_cols = {3, 6, 15}
    for rr in range(2, ws.max_row + 1):
        for c in amt_cols:
            ws.cell(row=rr, column=c).number_format = "#,##0.00"
        for c in date_cols:
            ws.cell(row=rr, column=c).number_format = "DD/MM/YYYY"

    _auto_fit_columns(ws, max_col=max_col)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _to_snake(name: str) -> str:
    s = str(name).strip().lower()
    out: list[str] = []
    prev_us = False
    for ch in s:
        if ch.isalnum():
            out.append(ch)
            prev_us = False
        else:
            if not prev_us:
                out.append("_")
                prev_us = True
    snake = "".join(out).strip("_")
    while "__" in snake:
        snake = snake.replace("__", "_")
    return snake


def normalize_gl_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize GL column names to canonical schema:
    invoice_number, date, customer, amount

    Rules:
    - Strip whitespace from all column names
    - Convert to lowercase snake_case
    - Map common variations to canonical names
    - Raise a clear error if any required column can't be matched
    """
    if df is None or df.empty:
        raise ValueError("GL file is empty or could not be read into a table.")

    out = df.copy()

    # (1) + (2) base normalization
    out.rename(columns={c: _to_snake(c) for c in out.columns}, inplace=True)
    out.columns = [_to_snake(c) for c in out.columns]

    # (3) synonym mapping (all values already snake_case)
    synonym_map: dict[str, set[str]] = {
        "invoice_number": {
            "invoice_number",
            "invoice_no",
            "invoice_num",
            "invoice",
            "inv_number",
            "inv_no",
            "inv_num",
            "inv",
        },
        "date": {
            "date",
            "invoice_date",
            "posting_date",
            "transaction_date",
            "gl_date",
        },
        "customer": {
            "customer",
            "customer_name",
            "client",
            "client_name",
            "debtor",
        },
        "amount": {
            "amount",
            "value",
            "total",
            "net_amount",
            "revenue",
            "gl_amount",
        },
    }

    reverse: dict[str, str] = {}
    for canonical, alts in synonym_map.items():
        for alt in alts:
            reverse[_to_snake(alt)] = canonical

    canonical_hits: dict[str, str] = {}
    for col in list(out.columns):
        key = _to_snake(col)
        if key not in reverse:
            continue
        canonical = reverse[key]
        if canonical in canonical_hits and canonical_hits[canonical] != col:
            raise ValueError(
                f"Ambiguous columns for '{canonical}': '{canonical_hits[canonical]}' and '{col}'. "
                "Please keep only one."
            )
        canonical_hits[canonical] = col

    # (4) required columns
    required = ["invoice_number", "date", "customer", "amount"]
    missing = [c for c in required if c not in canonical_hits]
    if missing:
        missing_list = ", ".join(missing)
        raise ValueError(
            f"Missing required column(s): {missing_list}. "
            "Please include columns that map to invoice_number, date, customer, amount."
        )

    out.rename(columns={src: canonical for canonical, src in canonical_hits.items()}, inplace=True)
    return out


def _pick_date_column(df: pd.DataFrame) -> str:
    # After normalization we always expect a canonical 'date' column.
    if "date" not in df.columns:
        raise ValueError(
            "Missing required column: date. Please include an invoice/transaction date column."
        )
    return "date"


def _target_threshold(performance_materiality: float, inherent_risk: str) -> float:
    risk_pct_by_level: dict[str, float] = {
        "significant": 0.10,
        "higher": 0.25,
        "lower": 0.50,
    }
    level = str(inherent_risk).strip().lower()
    if level not in risk_pct_by_level:
        raise ValueError(
            "Invalid inherent_risk. Expected one of: 'significant', 'higher', 'lower'."
        )
    return float(performance_materiality) * risk_pct_by_level[level]


def _fmt_title_case(value: str) -> str:
    v = str(value).strip()
    return v[:1].upper() + v[1:] if v else v


def _apply_table_style(ws, header_row: int, min_col: int, max_col: int) -> None:
    navy = PatternFill("solid", fgColor="0F2A52")
    white_bold = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D6DEEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(min_col, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = navy
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Borders for all populated cells
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            if cell.alignment is None or (cell.alignment.horizontal is None and cell.alignment.vertical is None):
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _auto_fit_columns(ws, min_col: int = 1, max_col: int | None = None) -> None:
    if max_col is None:
        max_col = ws.max_column
    for c in range(min_col, max_col + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c)].width = min(max(10, max_len + 2), 42)


def _write_header(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)


def _set_number_formats(ws, col_map: dict[str, int], start_row: int) -> None:
    amount_cols = {"Amount", "Total", "Total Value", "Threshold Used", "Interval", "Cumulative Value", "Target testing threshold"}
    date_cols = {"Date", "Cutoff Date", "Date/Time of run"}

    for name, idx in col_map.items():
        if name in amount_cols:
            for r in range(start_row, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "#,##0.00"
        if name in date_cols:
            for r in range(start_row, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "DD/MM/YYYY"


async def _read_gl_to_df(upload: UploadFile) -> pd.DataFrame:
    filename = (upload.filename or "").strip()
    ext = filename.split(".")[-1].lower() if "." in filename else ""
    data = await upload.read()
    if not data:
        raise ValueError("Uploaded file is empty.")

    bio = io.BytesIO(data)
    if ext == "csv":
        return pd.read_csv(bio)
    if ext in {"xls", "xlsx"}:
        return pd.read_excel(bio)

    # Fallback based on content type if extension is missing/misleading.
    ctype = (upload.content_type or "").lower()
    if "csv" in ctype:
        return pd.read_csv(bio)
    if "excel" in ctype or "spreadsheet" in ctype:
        return pd.read_excel(bio)

    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


@app.post("/upload")
async def upload(
    gl_file: UploadFile = File(...),
    performance_materiality: float = Form(...),
    inherent_risk: str = Form(...),
    control_risk: str = Form(...),
    sap_level: str = Form(...),
    confidence_level: int = Form(...),
    cutoff_date: dt.date = Form(...),
    test_negatives: str | None = Form(None),
    enable_target_testing: str | None = Form(None),
):
    """
    Accept GL + parameters, run tests, and return an Excel workpaper.
    """
    try:
        gl_df = normalize_gl_columns(await _read_gl_to_df(gl_file))

        # Clean key fields immediately after normalization (before any other processing).
        amt = gl_df["amount"].astype(str)
        amt = amt.str.strip()
        amt = amt.str.replace(r"(?i)\bAUD\b", "", regex=True)
        amt = amt.str.replace(r"[\$£€]", "", regex=True)
        amt = amt.str.replace(",", "", regex=False)
        gl_df["amount"] = pd.to_numeric(amt, errors="coerce")

        cust = gl_df["customer"].astype(str)
        cust = cust.str.replace(r"[\\\s]+$", "", regex=True)
        gl_df["customer"] = cust

        print("=== /upload: GL after normalize_gl_columns() + cleaning ===")
        print("Columns:", list(gl_df.columns))
        print("First 3 rows:\n", gl_df.head(3).to_string(index=False))
        print("Column dtypes:\n", gl_df.dtypes.astype(str).to_string())
        run_id = str(uuid4())
        run_dt = dt.datetime.now()
        enable_target = True if enable_target_testing is None else _coerce_bool(enable_target_testing)

        # Normalize population if negatives should be excluded.
        include_negatives = _coerce_bool(test_negatives)
        if not include_negatives:
            amounts = pd.to_numeric(gl_df["amount"], errors="coerce")
            gl_df = gl_df.loc[amounts.isna() | (amounts >= 0)].copy()

        # Normalize key columns for reporting
        date_column = _pick_date_column(gl_df)
        df = gl_df.copy()
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
        df[date_column] = pd.to_datetime(df[date_column], errors="coerce")

        invoice_col = "invoice_number"
        customer_col = "customer"

        if enable_target:
            target_df = target_testing(
                gl_transactions=df,
                performance_materiality=float(performance_materiality),
                risk_level=str(inherent_risk),
            )
            # Correct MUS methodology: remove target-tested items from population before MUS
            target_ids = (
                pd.Series(target_df.get(invoice_col, []))
                .dropna()
                .astype(str)
                .tolist()
            )
            target_id_set = set(target_ids)
            residual_df = df.loc[~df[invoice_col].astype(str).isin(target_id_set)].copy()
        else:
            target_df = df.iloc[0:0].copy()
            target_id_set = set()
            residual_df = df.copy()

        mus_df = mus_sampling(
            gl_transactions=residual_df,
            performance_materiality=float(performance_materiality),
            inherent_risk=str(inherent_risk),
            control_risk=str(control_risk),
            sap_level=str(sap_level),
            confidence_level=int(confidence_level),
            exclude_invoice_numbers=(target_id_set or None),
            invoice_col=invoice_col,
        )

        cutoff_df = cutoff_testing(
            gl_transactions=df,
            cutoff_date=cutoff_date,
            date_column=date_column,
        )

        # Compute parameter and population metrics
        pm = float(performance_materiality)
        pop_count = int(len(df))
        pop_value = float(df["amount"].sum(skipna=True))
        threshold_used = float(_target_threshold(pm, inherent_risk)) if enable_target else None
        target_tested_value = (
            float(pd.to_numeric(target_df.get("amount"), errors="coerce").sum(skipna=True))
            if enable_target
            else 0.0
        )
        residual_pop_value_for_mus = float(
            pd.to_numeric(residual_df.loc[residual_df["amount"] > 0, "amount"], errors="coerce").sum(skipna=True)
        )

        # MUS interval & cumulative values for sampled items
        population_mus = residual_df.loc[residual_df["amount"] > 0].copy()
        population_mus["__cume__"] = population_mus["amount"].cumsum()
        pop_value_pos = float(population_mus["amount"].sum(skipna=True))
        mus_params = mus_sample_size_parameters(
            population_value=pop_value_pos,
            performance_materiality=pm,
            inherent_risk=str(inherent_risk).strip().lower(),
            control_risk=str(control_risk).strip().lower(),
            sap_level=str(sap_level).strip().lower(),
        )
        ria_pct = float(mus_params["ria_pct"])
        confidence_factor = float(mus_params["confidence_factor"])
        sample_size = int(mus_params["sample_size"])
        interval = (pop_value_pos / sample_size) if sample_size > 0 else None

        # Build professional workbook with openpyxl
        wb = Workbook()
        wb.remove(wb.active)

        # -------- Sheet 1: Audit Parameters --------
        ws_params = wb.create_sheet("Audit Parameters")
        ws_params["A1"] = "Audit Parameters"
        ws_params["A1"].font = Font(bold=True, size=14, color="0F2A52")
        ws_params["A3"] = "Parameter"
        ws_params["B3"] = "Value"

        mus_formula_text = (
            "CEILING((Population Value × Confidence Factor) / Tolerable Misstatement)"
        )
        params_rows = [
            ("Performance Materiality", pm),
            ("Target Testing", "Enabled" if enable_target else "Disabled"),
            (
                "Target Testing Note",
                None
                if enable_target
                else "Target testing not performed - MUS applied to full population",
            ),
            ("Inherent Risk", _fmt_title_case(inherent_risk)),
            ("Control Risk", _fmt_title_case(control_risk)),
            ("SAP Level", _fmt_title_case(sap_level)),
            ("Confidence Level", int(confidence_level)),
            ("RIA (Risk of Incorrect Acceptance)", ria_pct / 100.0),
            ("Confidence Factor (Poisson, 0 expected misstatements)", confidence_factor),
            ("MUS Sample Size Formula", mus_formula_text),
            ("MUS Sample Size", sample_size),
            ("Cutoff Date", cutoff_date),
            ("Test Negatives", bool(include_negatives)),
            ("Run ID", run_id),
            ("Date/Time of run", run_dt),
            ("Total population count", pop_count),
            ("Total population value", pop_value),
            ("Less Target Tested Value", target_tested_value),
            ("Residual Population Value used for MUS (positive amounts)", residual_pop_value_for_mus),
            ("Target testing threshold", threshold_used),
        ]
        for i, (k, v) in enumerate(params_rows, start=4):
            ws_params.cell(row=i, column=1, value=k)
            ws_params.cell(row=i, column=2, value=v)

        _apply_table_style(ws_params, header_row=3, min_col=1, max_col=2)
        ws_params.freeze_panes = "A4"
        ws_params.column_dimensions["A"].width = 34
        ws_params.column_dimensions["B"].width = 34
        for r in range(4, 4 + len(params_rows)):
            k = ws_params.cell(row=r, column=1).value
            if k in {
                "Performance Materiality",
                "Total population value",
                "Less Target Tested Value",
                "Residual Population Value used for MUS (positive amounts)",
                "Target testing threshold",
            }:
                ws_params.cell(row=r, column=2).number_format = "#,##0.00"
            if k == "RIA (Risk of Incorrect Acceptance)":
                ws_params.cell(row=r, column=2).number_format = "0%"
            if k == "Cutoff Date":
                ws_params.cell(row=r, column=2).number_format = "DD/MM/YYYY"
            if k == "Date/Time of run":
                ws_params.cell(row=r, column=2).number_format = "DD/MM/YYYY HH:MM"

        # -------- Sheet 2: Target Testing --------
        ws_target = wb.create_sheet("Target Testing")
        target_headers = [
            "Invoice Number",
            "Date",
            "Customer",
            "Amount",
            "Threshold Used",
            "Reason Selected",
        ]
        if not enable_target:
            ws_target.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(target_headers))
            note_cell = ws_target.cell(
                row=1,
                column=1,
                value="Target testing not performed - MUS applied to full population",
            )
            note_cell.fill = PatternFill("solid", fgColor="0F2A52")
            note_cell.font = Font(bold=True, color="FFFFFF")
            note_cell.alignment = Alignment(horizontal="left", vertical="center")

            _write_header(ws_target, 2, target_headers)
            _apply_table_style(ws_target, header_row=2, min_col=1, max_col=len(target_headers))
            ws_target.freeze_panes = "A3"
            ws_target.auto_filter.ref = f"A2:{get_column_letter(len(target_headers))}2"
            _auto_fit_columns(ws_target)
        else:
            _write_header(ws_target, 1, target_headers)

            target_out = target_df.copy()
            target_out["__reason__"] = "Amount exceeds target testing threshold"
            for i, row in enumerate(target_out.itertuples(index=False), start=2):
                inv = getattr(row, invoice_col, None) if hasattr(row, invoice_col) else None
                dval = getattr(row, date_column, None) if hasattr(row, date_column) else None
                cust = getattr(row, customer_col, None) if hasattr(row, customer_col) else None
                amt = getattr(row, "amount", None)
                ws_target.cell(row=i, column=1, value=inv)
                ws_target.cell(row=i, column=2, value=(pd.to_datetime(dval).date() if pd.notna(dval) else None))
                ws_target.cell(row=i, column=3, value=cust)
                ws_target.cell(row=i, column=4, value=float(amt) if pd.notna(amt) else None)
                ws_target.cell(row=i, column=5, value=threshold_used)
                ws_target.cell(row=i, column=6, value="Amount exceeds threshold")

            # Summary row
            summary_row = ws_target.max_row + 2
            ws_target.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
            ws_target.cell(row=summary_row, column=2, value="Count").font = Font(bold=True)
            ws_target.cell(row=summary_row, column=3, value=int(len(target_out))).font = Font(bold=True)
            ws_target.cell(row=summary_row, column=4, value="Total value").font = Font(bold=True)
            total_target_value = float(pd.to_numeric(target_out.get("amount"), errors="coerce").sum(skipna=True))
            ws_target.cell(row=summary_row, column=5, value=total_target_value).font = Font(bold=True)
            ws_target.cell(row=summary_row, column=5).number_format = "#,##0.00"

            _apply_table_style(ws_target, header_row=1, min_col=1, max_col=len(target_headers))
            ws_target.freeze_panes = "A2"
            ws_target.auto_filter.ref = f"A1:{get_column_letter(len(target_headers))}1"
            # Formats
            for r in range(2, ws_target.max_row + 1):
                ws_target.cell(row=r, column=2).number_format = "DD/MM/YYYY"
                ws_target.cell(row=r, column=4).number_format = "#,##0.00"
                ws_target.cell(row=r, column=5).number_format = "#,##0.00"
            _auto_fit_columns(ws_target)

        # -------- Sheet 3: MUS Sample --------
        ws_mus = wb.create_sheet("MUS Sample")
        ws_mus.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        mus_note = (
            "MUS sample drawn from residual population (total less target-tested items)."
            if enable_target
            else "MUS sample drawn from full population (target testing not performed)."
        )
        note = ws_mus.cell(row=1, column=1, value=mus_note)
        note.fill = PatternFill("solid", fgColor="0F2A52")
        note.font = Font(bold=True, color="FFFFFF")
        note.alignment = Alignment(horizontal="left", vertical="center")
        mus_headers = [
            "Sample Number",
            "Invoice Number",
            "Date",
            "Customer",
            "Amount",
            "Cumulative Value",
            "Interval",
            "Reason Selected",
        ]
        _write_header(ws_mus, 2, mus_headers)

        mus_out = mus_df.copy()
        # Determine cumulative value in MUS population order
        mus_join = population_mus[[invoice_col, "__cume__"]].copy()
        mus_join.columns = [invoice_col, "__cume__"]
        if invoice_col in mus_out.columns:
            mus_out = mus_out.merge(mus_join, on=invoice_col, how="left")
        else:
            mus_out["__cume__"] = None

        for i, row in enumerate(mus_out.itertuples(index=False), start=3):
            inv = getattr(row, invoice_col, None) if hasattr(row, invoice_col) else None
            dval = getattr(row, date_column, None) if hasattr(row, date_column) else None
            cust = getattr(row, customer_col, None) if hasattr(row, customer_col) else None
            amt = getattr(row, "amount", None)
            cume = getattr(row, "__cume__", None) if hasattr(row, "__cume__") else None

            ws_mus.cell(row=i, column=1, value=i - 2)
            ws_mus.cell(row=i, column=2, value=inv)
            ws_mus.cell(row=i, column=3, value=(pd.to_datetime(dval).date() if pd.notna(dval) else None))
            ws_mus.cell(row=i, column=4, value=cust)
            ws_mus.cell(row=i, column=5, value=float(amt) if pd.notna(amt) else None)
            ws_mus.cell(row=i, column=6, value=float(cume) if pd.notna(cume) else None)
            ws_mus.cell(row=i, column=7, value=float(interval) if interval is not None else None)
            ws_mus.cell(row=i, column=8, value="Residual population MUS: cumulative value crossed sampling interval point")

        # Summary row
        summary_row = ws_mus.max_row + 2
        ws_mus.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
        ws_mus.cell(row=summary_row, column=2, value="Sample size").font = Font(bold=True)
        ws_mus.cell(row=summary_row, column=3, value=int(len(mus_out))).font = Font(bold=True)
        ws_mus.cell(row=summary_row, column=4, value="Total value").font = Font(bold=True)
        total_mus_value = float(pd.to_numeric(mus_out.get("amount"), errors="coerce").sum(skipna=True))
        ws_mus.cell(row=summary_row, column=5, value=total_mus_value).font = Font(bold=True)
        ws_mus.cell(row=summary_row, column=5).number_format = "#,##0.00"

        _apply_table_style(ws_mus, header_row=2, min_col=1, max_col=len(mus_headers))
        ws_mus.freeze_panes = "A3"
        ws_mus.auto_filter.ref = f"A2:{get_column_letter(len(mus_headers))}2"
        for r in range(2, ws_mus.max_row + 1):
            if r >= 3:
                ws_mus.cell(row=r, column=3).number_format = "DD/MM/YYYY"
            for c in [5, 6, 7]:
                if r >= 3:
                    ws_mus.cell(row=r, column=c).number_format = "#,##0.00"
        _auto_fit_columns(ws_mus)

        # -------- Sheet 4: Cutoff Testing --------
        ws_cutoff = wb.create_sheet("Cutoff Testing")
        cutoff_headers = [
            "Invoice Number",
            "Date",
            "Customer",
            "Amount",
            "Cutoff Position",
            "Days from Cutoff",
        ]

        # Split pre/post and add section headers
        cutoff_out = cutoff_df.copy()
        if date_column in cutoff_out.columns:
            cutoff_out[date_column] = pd.to_datetime(cutoff_out[date_column], errors="coerce")
        cutoff_out["__days__"] = None
        if date_column in cutoff_out.columns:
            cutoff_out["__days__"] = (cutoff_out[date_column].dt.date - cutoff_date).apply(
                lambda d: d.days if d is not None else None
            )

        navy = PatternFill("solid", fgColor="0F2A52")
        white_bold = Font(bold=True, color="FFFFFF")

        def section_title(row_idx: int, title: str) -> int:
            ws_cutoff.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(cutoff_headers))
            c = ws_cutoff.cell(row=row_idx, column=1, value=title)
            c.fill = navy
            c.font = white_bold
            c.alignment = Alignment(horizontal="left", vertical="center")
            return row_idx + 1

        r = 1
        r = section_title(r, f"Pre-cutoff transactions (before {cutoff_date.strftime('%d/%m/%Y')})")
        _write_header(ws_cutoff, r, cutoff_headers)
        _apply_table_style(ws_cutoff, header_row=r, min_col=1, max_col=len(cutoff_headers))
        r += 1

        pre_df = cutoff_out.loc[cutoff_out.get("cutoff_position") == "pre"].copy()
        for row in pre_df.itertuples(index=False):
            inv = getattr(row, invoice_col, None) if hasattr(row, invoice_col) else None
            dval = getattr(row, date_column, None) if hasattr(row, date_column) else None
            cust = getattr(row, customer_col, None) if hasattr(row, customer_col) else None
            amt = getattr(row, "amount", None)
            pos = getattr(row, "cutoff_position", None)
            days = getattr(row, "__days__", None)

            ws_cutoff.cell(row=r, column=1, value=inv)
            ws_cutoff.cell(row=r, column=2, value=(pd.to_datetime(dval).date() if pd.notna(dval) else None))
            ws_cutoff.cell(row=r, column=3, value=cust)
            ws_cutoff.cell(row=r, column=4, value=float(amt) if pd.notna(amt) else None)
            ws_cutoff.cell(row=r, column=5, value=pos)
            ws_cutoff.cell(row=r, column=6, value=int(days) if days is not None and pd.notna(days) else None)
            r += 1

        r += 1
        r = section_title(r, f"Post-cutoff transactions (after {cutoff_date.strftime('%d/%m/%Y')})")
        _write_header(ws_cutoff, r, cutoff_headers)
        _apply_table_style(ws_cutoff, header_row=r, min_col=1, max_col=len(cutoff_headers))
        r += 1

        post_df = cutoff_out.loc[cutoff_out.get("cutoff_position") == "post"].copy()
        for row in post_df.itertuples(index=False):
            inv = getattr(row, invoice_col, None) if hasattr(row, invoice_col) else None
            dval = getattr(row, date_column, None) if hasattr(row, date_column) else None
            cust = getattr(row, customer_col, None) if hasattr(row, customer_col) else None
            amt = getattr(row, "amount", None)
            pos = getattr(row, "cutoff_position", None)
            days = getattr(row, "__days__", None)

            ws_cutoff.cell(row=r, column=1, value=inv)
            ws_cutoff.cell(row=r, column=2, value=(pd.to_datetime(dval).date() if pd.notna(dval) else None))
            ws_cutoff.cell(row=r, column=3, value=cust)
            ws_cutoff.cell(row=r, column=4, value=float(amt) if pd.notna(amt) else None)
            ws_cutoff.cell(row=r, column=5, value=pos)
            ws_cutoff.cell(row=r, column=6, value=int(days) if days is not None and pd.notna(days) else None)
            r += 1

        ws_cutoff.freeze_panes = "A3"
        for rr in range(1, ws_cutoff.max_row + 1):
            ws_cutoff.cell(row=rr, column=2).number_format = "DD/MM/YYYY"
            ws_cutoff.cell(row=rr, column=4).number_format = "#,##0.00"
        _auto_fit_columns(ws_cutoff)

        # -------- Sheet 5: Population Summary --------
        ws_pop = wb.create_sheet("Population Summary")
        pop_headers = ["Invoice Number", "Date", "Customer", "Amount", "Selection"]
        _write_header(ws_pop, 1, pop_headers)

        # Keep normalized column names for processing; only use pretty headers in the worksheet.
        df_pop = df[[invoice_col, date_column, customer_col, "amount"]].copy()

        target_set = set(pd.Series(target_df.get(invoice_col, [])).dropna().astype(str).tolist())
        mus_set = set(pd.Series(mus_df.get(invoice_col, [])).dropna().astype(str).tolist())
        cutoff_set = set(pd.Series(cutoff_df.get(invoice_col, [])).dropna().astype(str).tolist())

        def selection_for(inv: str) -> str:
            tags = []
            if inv in target_set:
                tags.append("Target")
            if inv in mus_set:
                tags.append("MUS")
            if inv in cutoff_set:
                tags.append("Cutoff")
            return "/".join(tags) if tags else "Not Selected"

        for i, row in enumerate(df_pop.itertuples(index=False), start=2):
            inv = str(getattr(row, invoice_col))
            dval = getattr(row, date_column)
            cust = getattr(row, customer_col)
            amt = getattr(row, "amount")
            ws_pop.cell(row=i, column=1, value=inv)
            ws_pop.cell(row=i, column=2, value=(pd.to_datetime(dval).date() if pd.notna(dval) else None))
            ws_pop.cell(row=i, column=3, value=cust)
            ws_pop.cell(row=i, column=4, value=float(amt) if pd.notna(amt) else None)
            ws_pop.cell(row=i, column=5, value=selection_for(inv))

        _apply_table_style(ws_pop, header_row=1, min_col=1, max_col=len(pop_headers))
        ws_pop.freeze_panes = "A2"
        ws_pop.auto_filter.ref = f"A1:{get_column_letter(len(pop_headers))}1"
        for rr in range(2, ws_pop.max_row + 1):
            ws_pop.cell(row=rr, column=2).number_format = "DD/MM/YYYY"
            ws_pop.cell(row=rr, column=4).number_format = "#,##0.00"
        _auto_fit_columns(ws_pop)

        output = io.BytesIO()
        wb.save(output)

        # Persist a local copy so the workpaper exists even if browser download fails.
        OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
        disk_path = OUTPUTS_DIR / "Revenue_Test_Workpaper.xlsx"
        disk_path.write_bytes(output.getvalue())

        output.seek(0)
        out_name = "Revenue_Test_Workpaper.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{out_name}"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e)) from e


@app.get("/revenue-tests/target")
def run_target_testing() -> dict:
    return {
        "status": "not_implemented",
        "message": "target_testing now requires a GL DataFrame, performance_materiality, and risk_level. Wire this to an upload/ingestion endpoint next.",
    }


@app.get("/revenue-tests/mus")
def run_mus_sampling() -> dict:
    return {
        "status": "not_implemented",
        "message": "mus_sampling now requires a GL DataFrame plus PM/risk inputs. Wire this to an upload/ingestion endpoint next.",
    }


@app.get("/revenue-tests/cutoff")
def run_cutoff_testing() -> dict:
    return {
        "status": "not_implemented",
        "message": "cutoff_testing now requires a GL DataFrame, cutoff_date, and a date column. Wire this to the /upload endpoint.",
    }


@app.post("/test-documents")
async def test_documents(
    workpaper: UploadFile = File(...),
    invoices: list[UploadFile] = File(...),
    remittance_files: list[UploadFile] | None = File(None),
    bank_statements: list[UploadFile] = File(...),
) -> dict:
    """
    Accepts the workpaper Excel, invoice PDFs (batch), optional remittance advice PDFs (batch),
    and one or more bank statement PDFs.

    Invoices use Azure prebuilt-invoice only; remittances are uploaded via ``remittance_files`` and
    use the remittance OCR + OpenAI pipeline (no filename-based routing).

    Azure DI and OpenAI run concurrently (asyncio.gather / asyncio.to_thread): invoice PDFs,
    remittance PDFs, and bank statements are processed in parallel; bank transactions from every
    statement are merged into one pool for matching, with parsed remittances passed into
    ``_bank_match_for_sample`` unchanged.
    """
    try:
        workpaper_bytes = await workpaper.read()
        if not workpaper_bytes:
            raise ValueError("Workpaper file is empty.")
        sample_items = _load_sample_items_from_workpaper(workpaper_bytes)
        if not sample_items:
            raise ValueError(
                "No sample items found in workpaper. Ensure 'MUS Sample'/'Target Testing' include an 'Invoice Number' column."
            )

        invoice_files = invoices or []
        if len(invoice_files) == 0:
            raise ValueError("No invoice PDFs uploaded.")

        bank_files = bank_statements or []
        if len(bank_files) == 0:
            raise ValueError("No bank statement PDFs uploaded.")

        di_client = _azure_document_intelligence_client()
        oai = _openai_client()

        file_reads = await asyncio.gather(*[f.read() for f in invoice_files])
        bank_reads = await asyncio.gather(*[f.read() for f in bank_files])
        if not any(bank_reads):
            raise ValueError("All bank statement uploads are empty.")

        rem_batch = remittance_files if remittance_files else []
        rem_reads = await asyncio.gather(*[f.read() for f in rem_batch]) if rem_batch else []

        invoice_tasks = [
            _async_process_uploaded_pdf(di_client, oai, invoice_files[i].filename or "", file_reads[i])
            for i in range(len(invoice_files))
        ]
        remittance_tasks = [
            _async_process_remittance_pdf(di_client, oai, rem_batch[i].filename or "", rem_reads[i])
            for i in range(len(rem_batch))
        ]

        parsed_invoices: list[dict] = []
        parsed_remittances: list[dict] = []

        async def _gather_invoice_payloads() -> list[dict | None]:
            if not invoice_tasks:
                return []
            return await asyncio.gather(*invoice_tasks)

        async def _gather_remittance_payloads() -> list[dict | None]:
            if not remittance_tasks:
                return []
            return await asyncio.gather(*remittance_tasks)

        inv_results, rem_results, bank_txs = await asyncio.gather(
            _gather_invoice_payloads(),
            _gather_remittance_payloads(),
            _async_merge_bank_transactions(di_client, oai, list(bank_reads)),
        )
        for payload in inv_results:
            if payload is not None:
                parsed_invoices.append(payload)
        for payload in rem_results:
            if payload is not None:
                parsed_remittances.append(payload)

        if not parsed_invoices:
            raise ValueError(
                "No invoice PDFs could be parsed. Upload valid invoice PDFs (Azure prebuilt-invoice / OCR)."
            )

        results: list[dict] = []
        for item in sample_items:
            best_inv = _choose_best_invoice(item, parsed_invoices)

            # Fix invoice date confusion (YYYY-MM-DD where DD<=12 might actually be YYYY-DD-MM).
            if best_inv is not None:
                inv_dt = _norm_inv(best_inv.get("date") or "")
                gl_dt = _norm_inv(item.get("date") or "")
                m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", inv_dt)
                if m:
                    y, mm, dd = m.group(1), m.group(2), m.group(3)
                    try:
                        dd_i = int(dd)
                    except Exception:
                        dd_i = 99
                    if dd_i <= 12 and gl_dt:
                        cand_a = f"{y}-{mm}-{dd}"
                        cand_b = f"{y}-{dd}-{mm}"
                        try:
                            gl_parsed = pd.to_datetime(gl_dt, dayfirst=True, errors="coerce")
                            a_parsed = pd.to_datetime(cand_a, errors="coerce")
                            b_parsed = pd.to_datetime(cand_b, errors="coerce")
                            if (
                                pd.notna(gl_parsed)
                                and pd.notna(a_parsed)
                                and pd.notna(b_parsed)
                            ):
                                da = abs((a_parsed.date() - gl_parsed.date()).days)
                                db = abs((b_parsed.date() - gl_parsed.date()).days)
                                if db < da:
                                    best_inv["date"] = cand_b
                                    print(
                                        f"[DATE SWAP] inv={best_inv.get('invoice_number')} raw={cand_a} swapped={cand_b} gl={gl_dt} da={da} db={db}"
                                    )
                        except Exception:
                            pass

            gl_amt = item.get("gl_amount_ex_gst")
            inv_amt_ex = best_inv.get("amount_ex_gst") if best_inv else None
            variance = (
                round(float(inv_amt_ex) - float(gl_amt), 2)
                if (inv_amt_ex is not None and gl_amt is not None)
                else None
            )

            (
                bank_match,
                bank_date,
                conf,
                document_type,
                bank_desc,
                bank_cred,
            ) = _bank_match_for_sample(item, best_inv, bank_txs, parsed_remittances)

            overall = "Exception"
            if best_inv and inv_amt_ex is not None and gl_amt is not None:
                if abs(float(inv_amt_ex) - float(gl_amt)) <= 0.01 and bank_match in {"Yes", "Partial"}:
                    overall = "Pass"
                elif bank_match == "No":
                    overall = "Exception"
                else:
                    overall = "Exception"

            total_inc = best_inv.get("amount_inc_gst") if best_inv else None
            gst_component = _invoice_gst_component(best_inv, inv_amt_ex, total_inc)
            rcpt_vs_inv = _bank_receipt_vs_invoice_label(bank_match, bank_cred, total_inc)
            gst_check = bool(best_inv.get("gst_check_required")) if best_inv else False

            results.append(
                {
                    "invoice_number": item.get("invoice_number"),
                    "gl_amount_ex_gst": gl_amt,
                    "gl_date": (item.get("date") or ""),
                    "gl_customer": item.get("customer"),
                    "invoice_number_extracted": (best_inv.get("invoice_number") if best_inv else None),
                    "invoice_date_extracted": (_cell_date_iso(best_inv.get("date")) if best_inv else None),
                    "invoice_customer_extracted": (best_inv.get("customer") if best_inv else None),
                    "invoice_amount_ex_gst": inv_amt_ex,
                    "invoice_gst_amount": gst_component,
                    "invoice_total_inc_gst": total_inc,
                    "gst_check_required": gst_check,
                    "amount_variance": variance,
                    "line_items_summary": _line_items_summary(best_inv),
                    "bank_match": bank_match,
                    "bank_receipt_date": _cell_date_iso(bank_date),
                    "bank_description": bank_desc,
                    "bank_credit_amount": bank_cred,
                    "bank_receipt_vs_invoice": rcpt_vs_inv,
                    "document_type": document_type,
                    "match_confidence": conf,
                    "performance_obligation": "",
                    "overall_result": overall,
                    "auditor_notes": "",
                    "invoice_subtotal_ex_gst": inv_amt_ex,
                    "variance": variance,
                    "matched_invoice_source": (best_inv.get("source_filename") if best_inv else None),
                    "overridden": False,
                }
            )

        final_workpaper_bytes = _add_testing_results_sheet(workpaper_bytes, results)

        return {
            "results": results,
            "meta": {
                "sample_count": len(sample_items),
                "parsed_invoices": len(parsed_invoices),
                "parsed_remittances": len(parsed_remittances),
                "remittance_files_uploaded": len(rem_batch),
                "bank_transactions": len(bank_txs),
                "bank_statement_files": len([b for b in bank_reads if b]),
            },
            "final_workpaper": {
                "filename": (workpaper.filename or "Final_Tested_Workpaper.xlsx"),
                "content_base64": base64.b64encode(final_workpaper_bytes).decode("ascii"),
            },
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

